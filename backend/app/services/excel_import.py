"""Parsiranje Excel (.xlsx, .xlsm, .xls) u projekte i zadatke."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

import xlrd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _header_label(value: Any, index: int) -> str:
    text = _cell_str(value)
    return text if text else f"Polje {index + 1}"


def build_task_notes(headers: list[str], values: list[Any]) -> str | None:
    lines: list[str] = []
    for col_idx in range(1, len(values)):
        label = headers[col_idx] if col_idx < len(headers) else f"Polje {col_idx + 1}"
        text = _cell_str(values[col_idx] if col_idx < len(values) else None)
        if not text:
            continue
        lines.append(f"{label}: {text}")
    return "\n".join(lines) if lines else None


def parse_rows(rows: list[list[Any]]) -> tuple[list[dict], int]:
    """
    Vraća listu {title, notes} i broj preskočenih redova.
    Prvi red = zaglavlja; prva kolona = naziv zadatka.
    """
    if not rows:
        return [], 0

    header_row = rows[0]
    headers = [_header_label(h, i) for i, h in enumerate(header_row)]

    tasks: list[dict] = []
    skipped = 0

    for values in rows[1:]:
        if not any(v is not None and _cell_str(v) for v in values):
            skipped += 1
            continue
        title = _cell_str(values[0] if values else None)
        if not title:
            skipped += 1
            continue
        tasks.append({
            "title": title[:300],
            "notes": build_task_notes(headers, values),
        })

    return tasks, skipped


def _row_values_openpyxl(row) -> list[Any]:
    return [cell.value for cell in row]


def _parse_openpyxl_sheet(ws: Worksheet) -> tuple[list[dict], int]:
    rows = [_row_values_openpyxl(r) for r in ws.iter_rows()]
    return parse_rows(rows)


def _parse_xlrd_sheet(sheet: xlrd.sheet.Sheet, book: xlrd.book.Book) -> tuple[list[dict], int]:
    rows: list[list[Any]] = []
    for row_idx in range(sheet.nrows):
        values: list[Any] = []
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    values.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                except Exception:
                    values.append(cell.value)
            else:
                values.append(cell.value)
        rows.append(values)
    return parse_rows(rows)


def _parse_workbook_xlsx(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    results: list[dict] = []

    for ws in wb.worksheets:
        if ws.sheet_state and ws.sheet_state != "visible":
            continue
        name = (ws.title or "").strip()
        if not name:
            continue
        tasks, skipped = _parse_openpyxl_sheet(ws)
        if not tasks:
            continue
        results.append({
            "project_name": name[:200],
            "tasks": tasks,
            "skipped_rows": skipped,
        })

    wb.close()
    return results


def _parse_workbook_xls(file_bytes: bytes) -> list[dict]:
    book = xlrd.open_workbook(file_contents=file_bytes)
    results: list[dict] = []

    for idx in range(book.nsheets):
        sheet = book.sheet_by_index(idx)
        name = (sheet.name or "").strip()
        if not name:
            continue
        tasks, skipped = _parse_xlrd_sheet(sheet, book)
        if not tasks:
            continue
        results.append({
            "project_name": name[:200],
            "tasks": tasks,
            "skipped_rows": skipped,
        })

    return results


def _is_xls_format(file_bytes: bytes, filename: str | None) -> bool:
    name = (filename or "").lower()
    if name.endswith(".xls") and not name.endswith(".xlsx") and not name.endswith(".xlsm"):
        return True
    return len(file_bytes) >= 8 and file_bytes[:8] == XLS_MAGIC


def parse_workbook(file_bytes: bytes, filename: str | None = None) -> list[dict]:
    """Svaki sheet -> jedan projekat (naziv sheeta) + zadaci."""
    if _is_xls_format(file_bytes, filename):
        return _parse_workbook_xls(file_bytes)
    return _parse_workbook_xlsx(file_bytes)
